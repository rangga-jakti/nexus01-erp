"""
core/exports.py

Reusable export engine untuk Excel dan PDF.
Dipakai oleh semua modul — inventory, purchasing, sales, finance.

Kenapa di core?
Karena export logic tidak boleh duplikat di setiap modul.
Cukup panggil ExcelExporter atau PDFExporter dari mana saja.
"""

import io
from datetime import datetime
from django.http import HttpResponse


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL EXPORTER
# ─────────────────────────────────────────────────────────────────────────────

class ExcelExporter:
    """
    Export data ke Excel (.xlsx) pakai openpyxl.

    Contoh pemakaian:
        exp = ExcelExporter("Laporan Stok", "Inventory")
        exp.add_sheet("Stok", headers, rows)
        return exp.response("laporan_stok.xlsx")
    """

    # Warna tema Nexus (indigo)
    COLOR_HEADER_BG = "4F46E5"
    COLOR_HEADER_FG = "FFFFFF"
    COLOR_TITLE_BG  = "EEF2FF"
    COLOR_ALT_ROW   = "F8FAFC"
    COLOR_BORDER    = "E2E8F0"

    def __init__(self, title: str, company_name: str = "Nexus-01 ERP"):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        self.wb = Workbook()
        self.wb.remove(self.wb.active)   # hapus sheet default
        self.title = title
        self.company_name = company_name
        self._Font = Font
        self._Fill = PatternFill
        self._Align = Alignment
        self._Border = Border
        self._Side = Side

    def _header_font(self):
        return self._Font(name="Calibri", bold=True, color=self.COLOR_HEADER_FG, size=11)

    def _header_fill(self):
        return self._Fill(fill_type="solid", fgColor=self.COLOR_HEADER_BG)

    def _title_fill(self):
        return self._Fill(fill_type="solid", fgColor=self.COLOR_TITLE_BG)

    def _thin_border(self):
        side = self._Side(style="thin", color=self.COLOR_BORDER)
        return self._Border(left=side, right=side, top=side, bottom=side)

    def add_sheet(self, sheet_name: str, headers: list, rows: list,
                  col_widths: list = None):
        """
        Tambah sheet baru.

        headers : list of str — nama kolom
        rows    : list of list — baris data (urutan sama dengan headers)
        col_widths : list of int — lebar kolom opsional
        """
        ws = self.wb.create_sheet(title=sheet_name)

        # ── Title block ──────────────────────────────────────────────────────
        ws.merge_cells(f"A1:{self._col_letter(len(headers))}1")
        title_cell = ws["A1"]
        title_cell.value = self.company_name
        title_cell.font = self._Font(name="Calibri", bold=True, size=14,
                                     color="1E293B")
        title_cell.fill = self._title_fill()
        title_cell.alignment = self._Align(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 28

        ws.merge_cells(f"A2:{self._col_letter(len(headers))}2")
        sub_cell = ws["A2"]
        sub_cell.value = f"{self.title}  |  Diekspor: {datetime.now():%d %B %Y %H:%M}"
        sub_cell.font = self._Font(name="Calibri", size=10, color="64748B")
        sub_cell.fill = self._title_fill()
        sub_cell.alignment = self._Align(horizontal="left", vertical="center")
        ws.row_dimensions[2].height = 20

        ws.row_dimensions[3].height = 6  # spacer

        # ── Headers ──────────────────────────────────────────────────────────
        header_row = 4
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = self._header_font()
            cell.fill = self._header_fill()
            cell.alignment = self._Align(horizontal="center", vertical="center",
                                         wrap_text=True)
            cell.border = self._thin_border()
        ws.row_dimensions[header_row].height = 22

        # ── Data rows ────────────────────────────────────────────────────────
        for row_idx, row in enumerate(rows, start=header_row + 1):
            is_alt = (row_idx % 2 == 0)
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = self._Font(name="Calibri", size=10)
                cell.border = self._thin_border()
                cell.alignment = self._Align(vertical="center", wrap_text=False)
                if is_alt:
                    cell.fill = self._Fill(fill_type="solid",
                                           fgColor=self.COLOR_ALT_ROW)
                # Rata kanan untuk angka
                if isinstance(value, (int, float)):
                    cell.alignment = self._Align(horizontal="right",
                                                  vertical="center")
                    cell.number_format = '#,##0.##'
            ws.row_dimensions[row_idx].height = 18

        # ── Column widths ────────────────────────────────────────────────────
        if col_widths:
            for i, w in enumerate(col_widths, start=1):
                ws.column_dimensions[self._col_letter(i)].width = w
        else:
            # Auto-width berdasarkan konten
            for col_idx in range(1, len(headers) + 1):
                max_len = len(str(headers[col_idx - 1]))
                for row in rows:
                    if col_idx <= len(row):
                        max_len = max(max_len, len(str(row[col_idx - 1] or "")))
                ws.column_dimensions[self._col_letter(col_idx)].width = \
                    min(max_len + 4, 50)

        # Freeze panes di bawah header
        ws.freeze_panes = f"A{header_row + 1}"
        ws.auto_filter.ref = \
            f"A{header_row}:{self._col_letter(len(headers))}{header_row}"

        return ws

    def _col_letter(self, n: int) -> str:
        """Konversi nomor kolom ke huruf Excel (1→A, 26→Z, 27→AA, dst)."""
        result = ""
        while n:
            n, remainder = divmod(n - 1, 26)
            result = chr(65 + remainder) + result
        return result

    def response(self, filename: str) -> HttpResponse:
        """Return HttpResponse dengan file Excel."""
        buffer = io.BytesIO()
        self.wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument"
                          ".spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


# ─────────────────────────────────────────────────────────────────────────────
# PDF EXPORTER
# ─────────────────────────────────────────────────────────────────────────────

class PDFExporter:
    """
    Export data ke PDF pakai ReportLab.

    Contoh:
        pdf = PDFExporter("Laporan Purchase Order", company)
        pdf.add_table(headers, rows, col_widths)
        return pdf.response("po_report.pdf")
    """

    def __init__(self, title: str, company=None, landscape: bool = False):
        from reportlab.lib.pagesizes import A4, landscape as rl_landscape
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm

        self.buffer = io.BytesIO()
        self.title = title
        self.company = company
        self.landscape = landscape

        page_size = rl_landscape(A4) if landscape else A4
        self.page_size = page_size
        self.width, self.height = page_size

        self._colors = colors
        self._cm = cm
        self.story = []

        styles = getSampleStyleSheet()
        self.style_normal = styles['Normal']
        self.style_normal.fontName = 'Helvetica'
        self.style_normal.fontSize = 9

        self.style_title = ParagraphStyle(
            'NexusTitle',
            fontName='Helvetica-Bold',
            fontSize=16,
            textColor=colors.HexColor('#1E293B'),
            spaceAfter=4,
        )
        self.style_subtitle = ParagraphStyle(
            'NexusSub',
            fontName='Helvetica',
            fontSize=9,
            textColor=colors.HexColor('#64748B'),
            spaceAfter=16,
        )
        self.style_section = ParagraphStyle(
            'NexusSection',
            fontName='Helvetica-Bold',
            fontSize=10,
            textColor=colors.HexColor('#4F46E5'),
            spaceBefore=12,
            spaceAfter=4,
        )

    def build_header(self):
        from reportlab.platypus import Paragraph, Spacer, HRFlowable
        from reportlab.lib.units import cm

        company_name = self.company.name if self.company else "Nexus-01 ERP"
        self.story.append(Paragraph(company_name, self.style_title))
        self.story.append(Paragraph(
            f"{self.title}  ·  {datetime.now():%d %B %Y %H:%M}",
            self.style_subtitle
        ))
        self.story.append(HRFlowable(
            width="100%", thickness=2,
            color=self._colors.HexColor('#4F46E5'),
            spaceAfter=12,
        ))

    def add_info_grid(self, items: list):
        """
        Tambah grid info 2 kolom di bagian atas (label: value).
        items = [("Label", "Value"), ...]
        """
        from reportlab.platypus import Table, TableStyle
        from reportlab.lib.units import cm

        data = []
        for i in range(0, len(items), 2):
            row = []
            for item in items[i:i+2]:
                label, value = item
                row.extend([f"{label}:", str(value or "—")])
            while len(row) < 4:
                row.extend(["", ""])
            data.append(row)

        col_w = (self.width - 2 * self._cm) / 4
        t = Table(data, colWidths=[col_w * 0.4, col_w * 0.6] * 2)
        t.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
            ('TEXTCOLOR', (0, 0), (0, -1), self._colors.HexColor('#64748B')),
            ('TEXTCOLOR', (2, 0), (2, -1), self._colors.HexColor('#64748B')),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
        ]))
        self.story.append(t)
        from reportlab.platypus import Spacer
        self.story.append(Spacer(1, 12))

    def add_table(self, headers: list, rows: list,
                  col_widths: list = None, section_title: str = None):
        """Tambah tabel data ke PDF."""
        from reportlab.platypus import Table, TableStyle, Paragraph, Spacer

        if section_title:
            self.story.append(Paragraph(section_title, self.style_section))

        # Hitung lebar kolom
        usable_w = self.width - 2 * self._cm
        if col_widths:
            cw = [w * self._cm for w in col_widths]
        else:
            cw = [usable_w / len(headers)] * len(headers)

        # Build data
        data = [headers] + [[str(v) if v is not None else "—" for v in row]
                             for row in rows]

        t = Table(data, colWidths=cw, repeatRows=1)
        t.setStyle(TableStyle([
            # Header
            ('BACKGROUND', (0, 0), (-1, 0),
             self._colors.HexColor('#4F46E5')),
            ('TEXTCOLOR', (0, 0), (-1, 0),
             self._colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('TOPPADDING', (0, 0), (-1, 0), 7),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 7),

            # Data rows
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8.5),
            ('TOPPADDING', (0, 1), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 5),

            # Alternating rows
            ('ROWBACKGROUNDS', (0, 1), (-1, -1),
             [self._colors.white, self._colors.HexColor('#F8FAFC')]),

            # Grid
            ('GRID', (0, 0), (-1, -1),
             0.4, self._colors.HexColor('#E2E8F0')),
            ('LINEBELOW', (0, 0), (-1, 0),
             1, self._colors.HexColor('#4338CA')),
        ]))
        self.story.append(t)
        self.story.append(Spacer(1, 8))

    def add_summary(self, items: list):
        """Tambah baris summary/total di bawah tabel."""
        from reportlab.platypus import Table, TableStyle

        usable_w = self.width - 2 * self._cm
        data = [[item[0], item[1]] for item in items]
        t = Table(data, colWidths=[usable_w * 0.75, usable_w * 0.25])
        t.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('TEXTCOLOR', (0, 0), (-1, -1),
             self._colors.HexColor('#1E293B')),
            ('LINEABOVE', (0, 0), (-1, 0),
             1, self._colors.HexColor('#4F46E5')),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ]))
        self.story.append(t)

    def response(self, filename: str) -> HttpResponse:
        """Build PDF dan return HttpResponse."""
        from reportlab.platypus import SimpleDocTemplate
        from reportlab.lib.units import cm

        doc = SimpleDocTemplate(
            self.buffer,
            pagesize=self.page_size,
            rightMargin=self._cm,
            leftMargin=self._cm,
            topMargin=1.5 * self._cm,
            bottomMargin=1.5 * self._cm,
            title=self.title,
        )

        def footer(canvas, doc):
            canvas.saveState()
            canvas.setFont('Helvetica', 8)
            canvas.setFillColor(self._colors.HexColor('#94A3B8'))
            canvas.drawString(
                self._cm,
                0.7 * self._cm,
                f"Nexus-01 ERP  ·  {self.title}  ·  "
                f"Hal {doc.page}  ·  {datetime.now():%d/%m/%Y %H:%M}"
            )
            canvas.restoreState()

        doc.build(self.story, onFirstPage=footer, onLaterPages=footer)
        self.buffer.seek(0)

        response = HttpResponse(
            self.buffer.getvalue(),
            content_type="application/pdf"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response
